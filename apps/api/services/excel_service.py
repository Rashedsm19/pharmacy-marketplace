"""
The customer-facing spreadsheet.

A pharmacy with ten thousand lines is not going to type them into a form, so the
file is the onboarding path. It has to be obvious to fill in and forgiving to
read back: bilingual headers, an instructions sheet, a branch dropdown taken
from the pharmacy's own branches, and a reader that streams rather than loading
the workbook into memory.
"""
from __future__ import annotations

import csv
import io
import logging
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Iterator

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

logger = logging.getLogger("api.excel")

BRAND_TEAL = "FF0AA39B"
HEADER_TEXT = "FFFFFFFF"
NOTE_FILL = "FFF7F1E6"
GUIDE_FILL = "FFEFF6F5"
GUIDE_TEXT = "FF3F6B67"

# The template ships two guide rows above the customer's data: one explaining
# each column, one showing a filled example. Both begin with this marker in the
# first column, and the reader skips any row that carries it.
#
# This is not decoration. Before the marker existed, the example row sat at row 2
# and read_rows started at row 2 — so a customer who did the natural thing (type
# beneath the example) imported "Amoxicillin 500mg / BTH-2026-114 / 120 units" as
# real stock, and every weekly re-upload reset it back to 120 because the batch
# key matched. The marker is what makes the guide rows inert.
GUIDE_MARKER = "⟪"
GUIDE_NOTE = "⟪شرح⟫"
GUIDE_EXAMPLE = "⟪مثال⟫"

# Row 1 headers, row 2 explanation, row 3 example, row 4 onward is the customer.
FIRST_DATA_ROW = 4


@dataclass(frozen=True)
class Column:
    key: str
    header_ar: str
    header_en: str
    required: bool
    width: int
    help_ar: str
    example: str | int | float | None = None


# The order here is the order in the sheet, and the keys match BatchCreate.
COLUMNS: tuple[Column, ...] = (
    Column("product_name", "اسم الدواء", "Medicine name", True, 34,
           "الاسم كما هو لديك، عربي أو إنجليزي. يستخدم للمطابقة مع الكتالوج.",
           "Amoxicillin 500mg"),
    Column("barcode", "الباركود / GTIN", "Barcode / GTIN", False, 20,
           "أدق وسيلة للمطابقة. إن توفر فلا تتركه فارغا.", "6281000000017"),
    Column("sku", "كود المنتج لديك", "Your product code", False, 18,
           "كودك الداخلي. يبقى مرجعا لك ولا يشارك.", "MED-1042"),
    Column("batch_number", "رقم التشغيلة", "Batch number", True, 18,
           "مفتاح التحديث: رفع الملف مرة أخرى بالرقم نفسه يحدث الكمية ولا يكررها.",
           "BTH-2026-114"),
    Column("expiry_date", "تاريخ الانتهاء", "Expiry date", True, 16,
           "بصيغة YYYY-MM-DD أو كتاريخ Excel. ميلادي كما هو مطبوع على العلبة.",
           "2026-11-30"),
    Column("quantity", "الكمية", "Quantity", True, 12,
           "عدد صحيح أكبر من صفر.", 120),
    Column("unit_cost", "سعر التكلفة", "Unit cost", False, 14,
           "يستخدم في تقدير القيمة القابلة للاسترداد.", 12.5),
    Column("branch_name", "الفرع", "Branch", True, 24,
           "اختر من القائمة. الفروع مأخوذة من فروع منشأتك.", None),
    Column("supplier", "المورد", "Supplier", False, 22,
           "اختياري.", None),
    Column("purchase_order_number", "رقم أمر الشراء", "PO number", False, 18,
           "اختياري.", None),
    Column("requires_cold_chain", "يحتاج تبريد", "Cold chain", False, 14,
           "اكتب «نعم» أو «لا». الافتراضي «لا».", "لا"),
    Column("notes", "ملاحظات", "Notes", False, 30, "اختياري.", None),
)

REQUIRED_KEYS = tuple(c.key for c in COLUMNS if c.required)

# Accepted spellings for the yes/no column, in both languages.
TRUTHY = {"نعم", "yes", "y", "true", "1", "صح", "✔"}
FALSY = {"لا", "no", "n", "false", "0", ""}


def build_template(branch_names: list[str]) -> bytes:
    """The blank workbook a customer downloads."""
    workbook = Workbook()

    sheet = workbook.active
    sheet.title = "البيانات"
    sheet.sheet_view.rightToLeft = True

    header_font = Font(bold=True, color=HEADER_TEXT, size=11)
    header_fill = PatternFill("solid", fgColor=BRAND_TEAL)
    centre = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for index, column in enumerate(COLUMNS, start=1):
        label = f"{column.header_ar}{' *' if column.required else ''}\n{column.header_en}"
        cell = sheet.cell(row=1, column=index, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = centre
        sheet.column_dimensions[get_column_letter(index)].width = column.width
    sheet.row_dimensions[1].height = 34
    sheet.freeze_panes = f"A{FIRST_DATA_ROW}"

    # Row 2 — what to write in each column, so the answer is beside the question
    # rather than on another sheet.
    guide_fill = PatternFill("solid", fgColor=GUIDE_FILL)
    guide_font = Font(italic=True, size=10, color=GUIDE_TEXT)
    top_right = Alignment(horizontal="right", vertical="top", wrap_text=True)
    for index, column in enumerate(COLUMNS, start=1):
        duty = "إلزامي" if column.required else "اختياري"
        text = f"{duty} — {column.help_ar}"
        if index == 1:
            text = f"{GUIDE_NOTE} {text}"
        cell = sheet.cell(row=2, column=index, value=text)
        cell.fill = guide_fill
        cell.font = guide_font
        cell.alignment = top_right
    sheet.row_dimensions[2].height = 58

    # Row 3 — the same thing filled in, because a shape is easier to copy than
    # to infer from a description.
    example_fill = PatternFill("solid", fgColor=NOTE_FILL)
    for index, column in enumerate(COLUMNS, start=1):
        value = column.example
        if column.key == "branch_name" and branch_names:
            value = branch_names[0]
        if index == 1:
            value = f"{GUIDE_EXAMPLE} {value}"
        cell = sheet.cell(row=3, column=index, value=value)
        cell.fill = example_fill
        if column.key == "expiry_date":
            cell.number_format = "YYYY-MM-DD"

    # Dates typed into a text column are the most common import failure.
    date_column = get_column_letter(COLUMNS.index(next(c for c in COLUMNS if c.key == "expiry_date")) + 1)
    for row in range(3, 5002):
        sheet[f"{date_column}{row}"].number_format = "YYYY-MM-DD"

    if branch_names:
        branch_column = get_column_letter(
            COLUMNS.index(next(c for c in COLUMNS if c.key == "branch_name")) + 1
        )
        # Excel caps an inline list at 255 characters; fall back to free text.
        joined = ",".join(name.replace(",", " ") for name in branch_names)
        if len(joined) <= 250:
            validation = DataValidation(type="list", formula1=f'"{joined}"', allow_blank=False)
            validation.error = "اختر فرعا من القائمة"
            validation.errorTitle = "فرع غير معروف"
            sheet.add_data_validation(validation)
            # From the first real row: the guide rows are not the customer's to
            # validate, and the example already holds a valid branch.
            validation.add(f"{branch_column}{FIRST_DATA_ROW}:{branch_column}5001")

    _write_instructions(workbook, branch_names)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _write_instructions(workbook: Workbook, branch_names: list[str]) -> None:
    sheet = workbook.create_sheet("تعليمات")
    sheet.sheet_view.rightToLeft = True
    sheet.column_dimensions["A"].width = 26
    sheet.column_dimensions["B"].width = 12
    sheet.column_dimensions["C"].width = 76

    title = sheet.cell(row=1, column=1, value="كيف تملأ الملف")
    title.font = Font(bold=True, size=14, color="FF073F3C")

    lines = [
        "املأ ورقة «البيانات» صفا لكل تشغيلة — لا صفا لكل دواء.",
        "ابدأ الكتابة من الصف الرابع.",
        "الصفان الثاني والثالث شرح ومثال، ويتجاهلان تلقائيا عند الرفع — "
        "اتركهما أو احذفهما، لن يدخلا مخزونك.",
        "الأعمدة المعلمة بنجمة إلزامية، وما عداها اختياري.",
        "الصف الذي فيه خطأ لا يسقط الملف — يتخطى ويصلك في ملف الأخطاء بسببه ورقم سطره.",
        "رفع الملف مرة أخرى بنفس أرقام التشغيلات يحدث الكميات ولا يكررها.",
        "الدواء غير الموجود في كتالوج المنصة يضاف إلى مخزونك الخاص تلقائيا.",
    ]
    for offset, line in enumerate(lines, start=3):
        sheet.cell(row=offset, column=1, value=f"• {line}").alignment = Alignment(wrap_text=True)
        sheet.merge_cells(start_row=offset, start_column=1, end_row=offset, end_column=3)

    header_row = 3 + len(lines) + 1
    for index, label in enumerate(("العمود", "إلزامي", "الشرح"), start=1):
        cell = sheet.cell(row=header_row, column=index, value=label)
        cell.font = Font(bold=True, color=HEADER_TEXT)
        cell.fill = PatternFill("solid", fgColor=BRAND_TEAL)

    for offset, column in enumerate(COLUMNS, start=header_row + 1):
        sheet.cell(row=offset, column=1, value=column.header_ar)
        sheet.cell(row=offset, column=2, value="نعم" if column.required else "لا")
        note = sheet.cell(row=offset, column=3, value=column.help_ar)
        note.alignment = Alignment(wrap_text=True, vertical="top")

    if branch_names:
        row = header_row + len(COLUMNS) + 2
        sheet.cell(row=row, column=1, value="فروعك المتاحة").font = Font(bold=True)
        for offset, name in enumerate(branch_names, start=row + 1):
            sheet.cell(row=offset, column=1, value=name)


@dataclass
class ParsedRow:
    line_number: int
    values: dict[str, object] = field(default_factory=dict)


def _header_index(header_cells: list[object]) -> dict[str, int]:
    """Map our column keys onto whatever order the customer's file uses."""
    lookup: dict[str, str] = {}
    for column in COLUMNS:
        for spelling in (column.header_ar, column.header_en, column.key):
            lookup[_normalise_header(spelling)] = column.key

    mapping: dict[str, int] = {}
    for index, cell in enumerate(header_cells):
        if cell is None:
            continue
        # The template writes "Arabic *\nEnglish"; either line identifies it.
        for part in str(cell).replace("*", "").split("\n"):
            key = lookup.get(_normalise_header(part))
            if key and key not in mapping:
                mapping[key] = index
    return mapping


# Diacritics were removed from the product's Arabic, but a customer may still be
# holding a template downloaded before that. Folding them here means both
# spellings of a header map to the same column instead of the older file
# silently losing whichever column it named.
_HEADER_MARKS = re.compile(r"[\u064b-\u0652\u0670\u0640]")


def _normalise_header(text: str) -> str:
    stripped = _HEADER_MARKS.sub("", str(text))
    return "".join(stripped.split()).strip().lower()


def is_guide_row(raw) -> bool:
    """True for the template's own explanation and example rows.

    They are marked in the first cell rather than inferred from styling, because
    the workbook is read with `values_only=True` — fills and fonts are never
    loaded, so a colour could not be detected even if we wanted it to be. A real
    medicine name never opens with the marker.
    """
    if not raw:
        return False
    for cell in raw:
        if cell is None or str(cell).strip() == "":
            continue
        return str(cell).lstrip().startswith(GUIDE_MARKER)
    return False


def read_rows(content: bytes, filename: str) -> Iterator[ParsedRow]:
    """Stream the uploaded file row by row.

    Deliberately a generator: ten thousand rows must not be materialised at once
    on a small instance.
    """
    if filename.lower().endswith(".csv"):
        yield from _read_csv(content)
        return

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        sheet = workbook["البيانات"] if "البيانات" in workbook.sheetnames else workbook.worksheets[0]
        rows = sheet.iter_rows(values_only=True)
        try:
            header = list(next(rows))
        except StopIteration:
            return
        mapping = _header_index(header)
        if not mapping:
            raise ValueError("لم يتم التعرف على عناوين الأعمدة — استخدم القالب المرفق")

        for line_number, raw in enumerate(rows, start=2):
            if raw is None or all(cell is None or str(cell).strip() == "" for cell in raw):
                continue
            if is_guide_row(raw):
                continue
            values = {
                key: raw[index] if index < len(raw) else None
                for key, index in mapping.items()
            }
            yield ParsedRow(line_number=line_number, values=values)
    finally:
        workbook.close()


def _read_csv(content: bytes) -> Iterator[ParsedRow]:
    # Excel on Arabic Windows writes UTF-8 with a BOM more often than not.
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    try:
        header = next(reader)
    except StopIteration:
        return
    mapping = _header_index(list(header))
    if not mapping:
        raise ValueError("لم يتم التعرف على عناوين الأعمدة — استخدم القالب المرفق")

    for line_number, raw in enumerate(reader, start=2):
        if not any(str(cell).strip() for cell in raw):
            continue
        # Saving the template as CSV keeps the guide rows; they stay inert.
        if is_guide_row(raw):
            continue
        yield ParsedRow(
            line_number=line_number,
            values={
                key: raw[index] if index < len(raw) else None
                for key, index in mapping.items()
            },
        )


def coerce_date(value: object) -> date | None:
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    return None


def coerce_int(value: object) -> int | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        return int(float(str(value).strip().replace(",", "")))
    except (TypeError, ValueError):
        return None


def coerce_float(value: object) -> float | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        return float(str(value).strip().replace(",", ""))
    except (TypeError, ValueError):
        return None


def coerce_bool(value: object) -> bool:
    return str(value or "").strip().lower() in TRUTHY


def clean_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def build_errors_workbook(errors: list[dict]) -> bytes:
    """The rejected rows, with their line number and reason, ready to fix and resend."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "الأخطاء"
    sheet.sheet_view.rightToLeft = True

    headers = ("رقم السطر", "الدواء", "رقم التشغيلة", "سبب الرفض")
    widths = (12, 34, 20, 60)
    for index, (label, width) in enumerate(zip(headers, widths), start=1):
        cell = sheet.cell(row=1, column=index, value=label)
        cell.font = Font(bold=True, color=HEADER_TEXT)
        cell.fill = PatternFill("solid", fgColor="FFB4231F")
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"

    for offset, error in enumerate(errors, start=2):
        sheet.cell(row=offset, column=1, value=error.get("line"))
        sheet.cell(row=offset, column=2, value=error.get("product_name"))
        sheet.cell(row=offset, column=3, value=error.get("batch_number"))
        sheet.cell(row=offset, column=4, value=error.get("reason")).alignment = Alignment(
            wrap_text=True
        )

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()

"""
Importing a pharmacy's own stock from a spreadsheet.

These cover the promises the feature makes to a customer with ten thousand
medicines: the template round-trips, one bad row does not cost the file, a
second upload updates rather than duplicates, and nothing crosses between
pharmacies.
"""
from __future__ import annotations

import io
import uuid
from datetime import date, timedelta

import pytest
from openpyxl import load_workbook

from tests.conftest import auth, unique

XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def build_sheet(rows: list[dict], branch_name: str) -> bytes:
    """A customer's file, written the way their own export would write it."""
    from openpyxl import Workbook

    from services.excel_service import COLUMNS

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "البيانات"
    sheet.append([column.header_ar for column in COLUMNS])
    for row in rows:
        sheet.append(
            [
                row.get("product_name"),
                row.get("barcode"),
                row.get("sku"),
                row.get("batch_number"),
                row.get("expiry_date"),
                row.get("quantity"),
                row.get("unit_cost"),
                row.get("branch_name", branch_name),
                row.get("supplier"),
                row.get("purchase_order_number"),
                row.get("requires_cold_chain"),
                row.get("notes"),
            ]
        )
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


async def run_pending_imports() -> None:
    """Drive the worker directly instead of waiting on the scheduler."""
    from scheduler import process_import_jobs

    for _ in range(20):
        from database import AsyncSessionLocal
        from models.import_job import ImportJob, ImportStatus
        from sqlalchemy import func, select

        async with AsyncSessionLocal() as db:
            queued = await db.scalar(
                select(func.count(ImportJob.id)).where(
                    ImportJob.status == ImportStatus.QUEUED
                )
            )
        if not queued:
            return
        await process_import_jobs()


async def upload(client, token: str, content: bytes, name: str = "stock.xlsx") -> dict:
    response = await client.post(
        "/inventory/import",
        headers=auth(token),
        files={"file": (name, content, XLSX)},
    )
    assert response.status_code == 202, response.text
    await run_pending_imports()
    job_id = response.json()["id"]
    final = await client.get(f"/inventory/import/{job_id}", headers=auth(token))
    assert final.status_code == 200
    return final.json()


async def find_batches(client, token: str, batch_number: str) -> list[dict]:
    """Every batch with this number, paging because the list has no search."""
    found: list[dict] = []
    page = 1
    while True:
        response = await client.get(
            "/inventory/batches",
            headers=auth(token),
            params={"page": page, "page_size": 100},
        )
        assert response.status_code == 200, response.text
        body = response.json()
        found += [b for b in body["items"] if b["batch_number"] == batch_number]
        if page * 100 >= body["total"]:
            return found
        page += 1


async def a_branch_name(client, token: str) -> str:
    branches = await client.get("/branches", headers=auth(token))
    assert branches.status_code == 200
    items = branches.json()["items"]
    assert items, "the seeded pharmacy must have a branch"
    return items[0]["name_ar"] or items[0]["name"]


@pytest.mark.asyncio
async def test_the_template_downloads_and_opens(client, seller_token):
    response = await client.get("/inventory/import/template", headers=auth(seller_token))
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(XLSX)

    workbook = load_workbook(io.BytesIO(response.content))
    assert "البيانات" in workbook.sheetnames
    assert "تعليمات" in workbook.sheetnames

    headers = [cell.value for cell in workbook["البيانات"][1]]
    assert any("اسم الدواء" in str(h) for h in headers)
    assert any("تاريخ الانتهاء" in str(h) for h in headers)


@pytest.mark.asyncio
async def test_a_file_becomes_inventory(client, seller_token):
    branch = await a_branch_name(client, seller_token)
    code = unique("IMP")
    expiry = date.today() + timedelta(days=120)

    job = await upload(
        client,
        seller_token,
        build_sheet(
            [
                {
                    "product_name": f"دواء الاستيراد {code}",
                    "sku": code,
                    "batch_number": code,
                    "expiry_date": expiry.isoformat(),
                    "quantity": 40,
                    "unit_cost": 12.5,
                }
            ],
            branch,
        ),
    )

    assert job["status"] == "completed", job
    assert job["total_rows"] == 1
    assert job["created_batches"] == 1
    assert job["failed_rows"] == 0
    # Nothing in the catalogue matches an invented name, so it made its own.
    assert job["created_products"] == 1

    found = await find_batches(client, seller_token, code)
    assert found, "the imported batch must show in the pharmacy's inventory"
    assert found[0]["quantity"] == 40
    assert found[0]["expiry_date"] == expiry.isoformat()


@pytest.mark.asyncio
async def test_one_bad_row_does_not_cost_the_file(client, seller_token):
    branch = await a_branch_name(client, seller_token)
    code = unique("MIX")
    good = date.today() + timedelta(days=200)

    job = await upload(
        client,
        seller_token,
        build_sheet(
            [
                {
                    "product_name": f"سليم {code}",
                    "sku": f"{code}-A",
                    "batch_number": "OK-1",
                    "expiry_date": good.isoformat(),
                    "quantity": 10,
                },
                {  # no expiry date
                    "product_name": f"بلا تاريخ {code}",
                    "sku": f"{code}-B",
                    "batch_number": "BAD-1",
                    "quantity": 5,
                },
                {  # quantity is not a number
                    "product_name": f"كمية خاطئة {code}",
                    "sku": f"{code}-C",
                    "batch_number": "BAD-2",
                    "expiry_date": good.isoformat(),
                    "quantity": "كثير",
                },
                {  # branch the pharmacy does not have
                    "product_name": f"فرع مجهول {code}",
                    "sku": f"{code}-D",
                    "batch_number": "BAD-3",
                    "expiry_date": good.isoformat(),
                    "quantity": 7,
                    "branch_name": "فرع لا وجود له",
                },
                {
                    "product_name": f"سليم آخر {code}",
                    "sku": f"{code}-E",
                    "batch_number": "OK-2",
                    "expiry_date": good.isoformat(),
                    "quantity": 20,
                },
            ],
            branch,
        ),
    )

    assert job["status"] == "completed_with_errors"
    assert job["total_rows"] == 5
    assert job["created_batches"] == 2, "the good rows must still land"
    assert job["failed_rows"] == 3

    # Each rejection carries the line the customer must go and fix.
    lines = {error["line"] for error in job["errors"]}
    assert lines == {3, 4, 5}
    reasons = " ".join(error["reason"] for error in job["errors"])
    assert "تاريخ الانتهاء" in reasons
    assert "الكمية" in reasons
    assert "الفرع" in reasons

    # And the same rows come back as a sheet they can correct and re-upload.
    errors_file = await client.get(
        f"/inventory/import/{job['id']}/errors", headers=auth(seller_token)
    )
    assert errors_file.status_code == 200
    sheet = load_workbook(io.BytesIO(errors_file.content)).worksheets[0]
    assert sheet.max_row == 4  # header plus three rejected rows


@pytest.mark.asyncio
async def test_re_uploading_updates_instead_of_duplicating(client, seller_token):
    """The weekly sync case: same file, corrected quantities."""
    branch = await a_branch_name(client, seller_token)
    code = unique("SYNC")
    expiry = date.today() + timedelta(days=150)

    def sheet(quantity: int) -> bytes:
        return build_sheet(
            [
                {
                    "product_name": f"دواء المزامنة {code}",
                    "sku": code,
                    "batch_number": code,
                    "expiry_date": expiry.isoformat(),
                    "quantity": quantity,
                }
            ],
            branch,
        )

    first = await upload(client, seller_token, sheet(100))
    assert first["created_batches"] == 1
    assert first["created_products"] == 1

    second = await upload(client, seller_token, sheet(75))
    assert second["created_batches"] == 0, "a second upload must not duplicate stock"
    assert second["updated_batches"] == 1
    assert second["created_products"] == 0, "the product it made must be reused"
    assert second["matched_products"] == 1

    rows = await find_batches(client, seller_token, code)
    assert len(rows) == 1, "one batch, not two"
    assert rows[0]["quantity"] == 75, "the file is the source of truth"


@pytest.mark.asyncio
async def test_a_known_medicine_matches_the_catalogue(client, seller_token):
    """A name already in the catalogue must not spawn a private copy."""
    from database import AsyncSessionLocal
    from models.product import Product
    from sqlalchemy import select

    async with AsyncSessionLocal() as db:
        catalogue = (
            await db.execute(
                select(Product)
                .where(Product.owner_organization_id.is_(None))
                .limit(1)
            )
        ).scalar_one()
        known_name = catalogue.name_ar
        known_sku = catalogue.sku

    branch = await a_branch_name(client, seller_token)
    job = await upload(
        client,
        seller_token,
        build_sheet(
            [
                {
                    "product_name": known_name,
                    "batch_number": unique("CAT"),
                    "expiry_date": (date.today() + timedelta(days=300)).isoformat(),
                    "quantity": 12,
                }
            ],
            branch,
        ),
    )

    assert job["matched_products"] == 1
    assert job["created_products"] == 0, "it must reuse the catalogue entry"

    batches = await client.get(
        "/inventory/batches", headers=auth(seller_token), params={"page_size": 100}
    )
    linked = [b for b in batches.json()["items"] if b["product_sku"] == known_sku]
    assert linked, "the batch must point at the catalogue product"


@pytest.mark.asyncio
async def test_an_import_is_invisible_to_another_pharmacy(
    client, seller_token, buyer_token, admin_token
):
    """The isolation guarantee, end to end."""
    branch = await a_branch_name(client, seller_token)
    code = unique("ISO")

    job = await upload(
        client,
        seller_token,
        build_sheet(
            [
                {
                    "product_name": f"دواء معزول {code}",
                    "sku": code,
                    "batch_number": code,
                    "expiry_date": (date.today() + timedelta(days=90)).isoformat(),
                    "quantity": 33,
                }
            ],
            branch,
        ),
    )
    assert job["created_batches"] == 1
    assert await find_batches(client, seller_token, code), "the owner must see it"

    # The other pharmacy sees neither the stock nor the product it created.
    assert not await find_batches(client, buyer_token, code), "the stock leaked"

    their_products = await client.get(
        "/products", headers=auth(buyer_token), params={"search": code}
    )
    assert code not in {p["sku"] for p in their_products.json()["items"]}

    # Nor the job itself — and the id must not confirm it exists.
    their_job = await client.get(
        f"/inventory/import/{job['id']}", headers=auth(buyer_token)
    )
    assert their_job.status_code == 404

    # The platform admin sees it.
    admin_view = await client.get(
        f"/inventory/import/{job['id']}", headers=auth(admin_token)
    )
    assert admin_view.status_code == 200


@pytest.mark.asyncio
async def test_the_job_list_is_scoped_to_the_pharmacy(client, seller_token, buyer_token):
    branch = await a_branch_name(client, seller_token)
    code = unique("قائمة")
    job = await upload(
        client,
        seller_token,
        build_sheet(
            [
                {
                    "product_name": code,
                    "batch_number": code,
                    "expiry_date": (date.today() + timedelta(days=60)).isoformat(),
                    "quantity": 3,
                }
            ],
            branch,
        ),
    )

    seller_org = (
        await client.get("/organizations/me", headers=auth(seller_token))
    ).json()["id"]

    mine = await client.get("/inventory/import", headers=auth(seller_token))
    assert mine.status_code == 200
    assert job["id"] in {item["id"] for item in mine.json()["items"]}
    assert all(item["organization_id"] == seller_org for item in mine.json()["items"])

    theirs = await client.get("/inventory/import", headers=auth(buyer_token))
    assert theirs.status_code == 200
    assert job["id"] not in {item["id"] for item in theirs.json()["items"]}
    assert all(
        item["organization_id"] != seller_org for item in theirs.json()["items"]
    ), "another pharmacy's imports leaked into the list"


@pytest.mark.asyncio
async def test_a_csv_is_accepted_too(client, seller_token):
    branch = await a_branch_name(client, seller_token)
    code = unique("CSV")
    expiry = (date.today() + timedelta(days=45)).isoformat()
    csv_text = (
        "اسم الدواء,الباركود / GTIN,كود المنتج لديك,رقم التشغيلة,تاريخ الانتهاء,"
        "الكمية,سعر التكلفة,الفرع,المورّد,رقم أمر الشراء,يحتاج تبريد,ملاحظات\n"
        f"دواء CSV {code},,{code},CSV-1,{expiry},9,4.5,{branch},,,,\n"
    )
    job = await upload(
        client, seller_token, csv_text.encode("utf-8-sig"), name="stock.csv"
    )
    assert job["status"] == "completed", job
    assert job["created_batches"] == 1


@pytest.mark.asyncio
async def test_a_file_that_is_not_a_spreadsheet_is_refused(client, seller_token):
    response = await client.post(
        "/inventory/import",
        headers=auth(seller_token),
        files={"file": ("notes.txt", b"hello", "text/plain")},
    )
    assert response.status_code == 415

    disguised = await client.post(
        "/inventory/import",
        headers=auth(seller_token),
        files={"file": ("stock.xlsx", b"not really a workbook", XLSX)},
    )
    assert disguised.status_code == 400


@pytest.mark.asyncio
async def test_capacity_is_reported_before_uploading(client, seller_token):
    response = await client.get("/inventory/import/capacity", headers=auth(seller_token))
    assert response.status_code == 200
    body = response.json()
    assert body["limit"] == 10000
    assert body["remaining"] == body["limit"] - body["used"]


@pytest.mark.asyncio
async def test_a_missing_job_is_not_found(client, seller_token):
    response = await client.get(
        f"/inventory/import/{uuid.uuid4()}", headers=auth(seller_token)
    )
    assert response.status_code == 404


@pytest.mark.asyncio
async def test_the_platform_admin_can_still_download_the_template(client, admin_token):
    """Regression: the admin got 403 and the screen said "تعذّر تحميل القالب".

    An admin has no pharmacy, but the template is a blank form — only the branch
    dropdown needs an organization.
    """
    response = await client.get("/inventory/import/template", headers=auth(admin_token))
    assert response.status_code == 200, response.text
    assert response.headers["content-type"].startswith(XLSX)

    workbook = load_workbook(io.BytesIO(response.content))
    assert "البيانات" in workbook.sheetnames
    assert "تعليمات" in workbook.sheetnames


@pytest.mark.asyncio
async def test_capacity_explains_rather_than_refuses_for_an_admin(client, admin_token):
    """The screen needs an answer it can explain, not a failed request."""
    response = await client.get("/inventory/import/capacity", headers=auth(admin_token))
    assert response.status_code == 200
    assert response.json()["can_import"] is False

    listed = await client.get("/inventory/import", headers=auth(admin_token))
    assert listed.status_code == 200
    assert listed.json()["total"] == 0


@pytest.mark.asyncio
async def test_a_pharmacy_is_told_it_can_import(client, seller_token):
    response = await client.get("/inventory/import/capacity", headers=auth(seller_token))
    assert response.status_code == 200
    assert response.json()["can_import"] is True


@pytest.mark.asyncio
async def test_an_admin_uploading_is_told_where_imports_belong(client, admin_token):
    refused = await client.post(
        "/inventory/import",
        headers=auth(admin_token),
        files={"file": ("x.xlsx", b"PK\x03\x04ignored", XLSX)},
    )
    assert refused.status_code == 403
    assert "حساب المنشأة" in refused.json()["detail"]


@pytest.mark.asyncio
async def test_the_downloaded_template_imports_nothing_by_itself(client, seller_token):
    """Regression, and the worst bug this feature had.

    The template used to carry its example at row 2 while the reader started at
    row 2. A customer who typed beneath the example — the natural thing to do —
    imported "Amoxicillin 500mg / BTH-2026-114 / 120" as real stock, and every
    weekly re-upload reset it to 120 again because the batch key matched.
    """
    from services.excel_service import read_rows

    response = await client.get("/inventory/import/template", headers=auth(seller_token))
    assert response.status_code == 200

    rows = list(read_rows(response.content, "template.xlsx"))
    assert rows == [], f"the blank template must import nothing, got {len(rows)} row(s)"


@pytest.mark.asyncio
async def test_the_template_explains_every_column_before_the_example(client, seller_token):
    from services.excel_service import COLUMNS, GUIDE_EXAMPLE, GUIDE_NOTE

    response = await client.get("/inventory/import/template", headers=auth(seller_token))
    sheet = load_workbook(io.BytesIO(response.content))["البيانات"]

    guide = [sheet.cell(row=2, column=i).value for i in range(1, len(COLUMNS) + 1)]
    assert str(guide[0]).startswith(GUIDE_NOTE)
    assert len([g for g in guide if g]) == len(COLUMNS), "every column must be explained"
    # Each one says whether it is required, in the customer's language.
    assert all(("إلزامي" in str(g)) or ("اختياري" in str(g)) for g in guide)

    example = [sheet.cell(row=3, column=i).value for i in range(1, len(COLUMNS) + 1)]
    assert str(example[0]).startswith(GUIDE_EXAMPLE)
    assert example[4], "the example must show a real expiry date"


@pytest.mark.asyncio
async def test_a_customer_filling_the_template_imports_only_their_rows(
    client, seller_token
):
    """The whole point: guide rows stay, the customer's row lands, once."""
    from openpyxl import load_workbook as open_wb

    from services.excel_service import FIRST_DATA_ROW

    branch = await a_branch_name(client, seller_token)
    code = unique("TPL")

    downloaded = await client.get(
        "/inventory/import/template", headers=auth(seller_token)
    )
    workbook = open_wb(io.BytesIO(downloaded.content))
    sheet = workbook["البيانات"]
    for index, value in enumerate(
        [
            f"دواء القالب {code}", None, code, code,
            (date.today() + timedelta(days=95)).isoformat(), 33, 7.5, branch,
            None, None, None, None,
        ],
        start=1,
    ):
        sheet.cell(row=FIRST_DATA_ROW, column=index, value=value)

    buffer = io.BytesIO()
    workbook.save(buffer)
    job = await upload(client, seller_token, buffer.getvalue(), name="filled.xlsx")

    assert job["total_rows"] == 1, "the guide rows must not be counted as data"
    assert job["created_batches"] == 1
    assert job["failed_rows"] == 0

    rows = await find_batches(client, seller_token, code)
    assert len(rows) == 1
    assert rows[0]["quantity"] == 33

    # And nothing resembling the shipped example got in.
    phantom = await find_batches(client, seller_token, "BTH-2026-114")
    assert phantom == [], "the template's example leaked into inventory"
